"""
This script is used to run a box trial task in the PsychoPy application.
"""

from psychopy import core, visual, event, data, clock, gui
import glob
from random import shuffle
import Constants
from Trial import Trial
from Form import Form
import pandas
from collections import defaultdict, namedtuple
import os
from InfoScene import InfoScene
from openpyxl import load_workbook, Workbook

def build_location_sequence(sequence):
    """
    Builds a random sequence if the sequence is none. 
    Else it parses the sequence string
    Params:
        sequence: A sequence string or none
    Returns:
        The sequence as a list of int
    """
    if(sequence is None):
        random_loc = [x for x in range(1, 13)]
        shuffle(random_loc)
        return random_loc
    return [int(x) for x in sequence.split(",")]


class ExperimentManager(object):
    """
    This class is a high level manager class using a state manager pattern.
    """
    def __init__(self, win, mouse, timer, writer, summary_output):
        """
        Initializes the experiment class
        Params:
            win: An instance representing a window on screen
            mouse: An instance representing the mouse
            timer: A clock instance to be used for timing
            writer: An instance of a pandas writer
            summary_output: A defaultdict for storing summary experiment output
        Returns:
            None
        """
        self.win = win
        self.mouse = mouse
        self.timer = timer
        self.block_count = 0
        self.trial_count = 0
        self.question_count = 0
        self.is_running = True
        self.writer = writer
        self.summary_output = summary_output
        self.practice_run() # Start the practice run NOTE: Only one practice run

    def practice_run(self):
        """
        This method creates a new practice trial by using the data specified in Constants
        Params:
            None
        Returns:
            None
        """
        practice_data = data.importConditions(Constants.PRACTICE_RUN)
        self.handler = data.TrialHandler(practice_data, 1, method="sequential")
        self.create_trial(self.handler.next(), Constants.PRACTICE_TRIAL_INFO)


    def next_block(self):
        """
        Goes to the next block of trials
        Params:
            None
        Returns:
            None
        """
        if self.block_count >= len(Constants.BLOCK_FILES): # if we have run out of blocks
            self.is_running = False
            return
        block_data = data.importConditions(Constants.BLOCK_FILES[self.block_count])
        self.block_count += 1
        self.handler = data.TrialHandler(block_data, 1, method="sequential")
        self.next_trial()

    def create_trial(self, trial_data, text_override=None):
        """
        Parses trial data and creates a new trial
        Params:
            trial_data: Trial conditions from a TrialHandler
            text_override: Text override for the trial info scene
        Returns:
            None
        """
        self.trial_count += 1
        #Parsing data from the data frame
        colours = (trial_data["Colour0"], trial_data["Colour1"], trial_data["ColourName0"], trial_data["ColourName1"])
        sequence = trial_data["Sequence"][1:-1] #The sequence string is specified with quotes marking beginning and end. The list slice will remove these.
        location_sequence = build_location_sequence(trial_data["Location_Sequence"])
        #Creating and running the trial
        trial_output = defaultdict(list) #The default dict allows for dynamically adding keywords with empty lists as the default value
        Trial(self.win, colours, sequence, self.mouse, trial_output, self.timer, location_sequence, self, text_override) # The trial object is created with each sequence. It is responsible for outputting necessary data from the trial

    def completed_trial(self, failed):
        """
        Sets a boolean flag and continues to next trial
        Params:
            failed: A boolean flag to set if the trial failed
        Returns:
            None
        """
        self.save_trial()
        self.failed_last = failed
        self.next_trial()

    def save_trial(self):
        """
        Saves the trial data
        Params:
            None
        Returns:
            None
        """
        data = self.scene.save()
        self.summary_output[f"Box_Num_{self.trial_count}"].append(data["Box_Num"][-1])
        self.summary_output[f"Probability_Estimate_{self.trial_count}"].append(data["Probability_Estimates"][-1])
        self.summary_output[f"Decision{self.trial_count}"].append(data["Decision"][-1])
        pandas.DataFrame(data).to_excel(self.writer, sheet_name=f"block{self.block_count}_trial{self.trial_count-1}")

    def next_trial(self):
        """
        Continues to the next trial in handler
        Params:
            None
        Returns:
            None
        """
        try: # try until handler is empty
            trial_data = self.handler.next()
        except StopIteration: # if handler is empty run form
            self.run_form()
            return
        if(self.trial_count == 0):
            if self.failed_last == True:
                self.create_trial(trial_data, Constants.FAILED_PRACTICE_TRIAL)
            else:
                self.create_trial(trial_data, Constants.COMPLETED_PRACTICE_TRIAL)
            return
        self.create_trial(trial_data)

    def to_trial(self):
        """
        A simple python hack to use as a callback function to proceed to next block
        """
        self.next_block()

    def form_ended(self, form_output):
        """
        Callback for a form class to save data and continue to next block
        Params:
            form_output: defaultdict to save the data into
        Returns:
            None
        """
        for item in form_output["Answer"]:
            self.summary_output[f"{Constants.FORM_FILES[self.block_count]}_Question{self.question_count}"] = item
            self.question_count += 1
        pandas.DataFrame(form_output).to_excel(self.writer, sheet_name=f"{Constants.FORM_FILES[self.block_count]}_answered")
        self.next_block()


    def run_form(self):
        """
        Runs a form
        Params:
            None
        Returns:
            None
        """
        if(Constants.FORM_FILES[self.block_count] is None):
            self.next_block()
            return
        #Parsing questionnaire data and creating form
        data_frame = pandas.read_excel(Constants.FORM_FILES[self.block_count])
        form_output = defaultdict(list) #Output stream
        Form(self.win, data_frame, form_output, self.timer, self)

    def update(self):
        """
        Updates the manager state.
        Params:
            None
        Returns:
            None
        """
        self.scene.check_input()
        self.scene.draw()
        self.win.flip()

def get_subject_info():
    """
    Gets the subject info through a dialogue box
    Params:
        None
    Returns:
        list of info or raises a valueError
    """
    dlg = gui.Dlg(title="Box Task Experiment")
    dlg.addText("Subject info")
    dlg.addField("ID")
    dlg.addField("Gender", choices=["Male", "Female"])
    dlg.addField("Age")
    data = dlg.show()
    if(dlg.OK):
        return data
    else:
        raise ValueError("Participant did not fill out the dialogue box")

def main():
    """
    Main program function to start the experiment manager an initializes needed dependencies
    Params:
        None
    Returns:
        None
    """
    subject_data = get_subject_info()
    summary_data = defaultdict(list)
    summary_data["ID"].append(subject_data[0])
    summary_data["Sex"].append(subject_data[1])
    summary_data["Age"].append(subject_data[2])
    with pandas.ExcelWriter(f"./data/ID_{subject_data[0]}.xlsx") as writer:
        win = visual.Window(Constants.WINDOW_SIZE, units="pix", fullscr=Constants.FULLSCREEN, color=Constants.BACKGROUND_COLOR); # NOTE: pixel units are not scalable.
        mouse = event.Mouse()
        timer = clock.Clock()
        manager = ExperimentManager(win, mouse, timer, writer, summary_data)
        while manager.is_running:
            manager.update()
        print(summary_data)
        try:
            book = load_workbook('./data/Summary.xlsx')
            write_headers = False
        except FileNotFoundError:
            book = Workbook()
            book.create_sheet("Main")
            write_headers = True
        summary = pandas.ExcelWriter("./data/Summary.xlsx", engine="openpyxl", mode="a")
        summary.book = book
        summary.sheets = dict((ws.title, ws) for ws in book.worksheets)
        append_row = book["Main"].max_row
        pandas.DataFrame(summary_data).to_excel(summary, sheet_name="Main", startrow=append_row, header=write_headers)
        summary.save()

#Start program  
if __name__ == "__main__":
    main()
